"""
Architecture Floor Plan — Sign Takeoff & PDF Annotator
=======================================================
Extracts all rooms/spaces from an architectural PDF floor plan,
generates an Excel sign takeoff sheet, and overlays color-coded
dot markers on the original PDF.

Tested on: Union at Tower District — 4th Floor Plan (Sheet A1.4)
Architect: BVH Architecture | Date: 03-28-2025

Dependencies:
    pip install pdfplumber pypdf reportlab openpyxl

Usage:
    python sign_takeoff_annotator.py

Key Problem Solved — PDF Rotation + Coordinate Transform:
----------------------------------------------------------
Architectural PDFs are frequently stored in portrait orientation
(raw MediaBox) with a /Rotate metadata flag that tells the viewer
to display them in landscape. This creates a mismatch between:

  1. pdfplumber coordinate space  — works in the DISPLAY (rotated) space.
                                     Origin = top-left, y increases downward.
                                     Reports page as e.g. 3024 wide x 2160 tall.

  2. ReportLab canvas space       — draws in the RAW (unrotated) space.
                                     Origin = bottom-left, y increases upward.
                                     Must use the raw MediaBox e.g. 2160 wide x 3024 tall.

For a page with /Rotate = 90 (viewer rotates 90° CCW to display):
  The transform from pdfplumber coords → raw canvas coords is:

        canvas_x = plumber_y          ← NOT (raw_width  - plumber_y)
        canvas_y = plumber_x          ← NOT (raw_height - plumber_x)

  Common mistakes:
    ✗  canvas_x = raw_width  - plumber_y   (shifts everything, wrong direction)
    ✗  canvas_x = plumber_x               (ignores rotation entirely)
    ✓  canvas_x = plumber_y
    ✓  canvas_y = plumber_x

  Derivation:
    For /Rotate=90, the viewer applies a 90° CCW rotation to the raw page.
    In that transform: display_x = raw_y, display_y_bl = raw_width - raw_x
    pdfplumber's y is from the top, so: plumber_y = display_height - display_y_bl
    Solving back:
        raw_x = display_height - plumber_y ... simplifies to raw_x = plumber_y
                (because display_height == raw_width for 90° rotation)
        raw_y = display_x = plumber_x
    ReportLab canvas origin is bottom-left of the raw page, so:
        canvas_x = raw_x = plumber_y
        canvas_y = raw_y = plumber_x

How to detect this situation:
    from pypdf import PdfReader
    reader = PdfReader("plan.pdf")
    rotation = reader.pages[0].get('/Rotate', 0)   # 0, 90, 180, or 270
    mediabox = reader.pages[0].mediabox
    raw_w = float(mediabox.right  - mediabox.left)
    raw_h = float(mediabox.top    - mediabox.bottom)
    # pdfplumber will report width=raw_h, height=raw_w when rotation=90 or 270

Generalised transforms for all rotation values:
    /Rotate = 0:    canvas_x = plumber_x,           canvas_y = raw_h - plumber_y
    /Rotate = 90:   canvas_x = plumber_y,            canvas_y = plumber_x
    /Rotate = 180:  canvas_x = raw_w - plumber_x,   canvas_y = plumber_y
    /Rotate = 270:  canvas_x = raw_h - plumber_y,   canvas_y = raw_w - plumber_x
"""

import re
import io
import pdfplumber
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.colors import HexColor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────

INPUT_PDF  = "4th_Floor_Union_at_Tower_Dist.pdf"
OUTPUT_PDF = "4th_Floor_Annotated_Signs.pdf"
OUTPUT_XLS = "sign_takeoff.xlsx"

# ── Step 1: Detect page rotation and set up coordinate transform ───────────────

def get_page_info(pdf_path):
    """Return raw MediaBox dimensions and rotation from the PDF."""
    reader = PdfReader(pdf_path)
    page = reader.pages[0]
    rotation = int(page.get('/Rotate', 0))
    mb = page.mediabox
    raw_w = float(mb.right  - mb.left)
    raw_h = float(mb.top    - mb.bottom)
    return raw_w, raw_h, rotation


def make_canvas_transform(raw_w, raw_h, rotation):
    """
    Return a function that converts pdfplumber (px, py) → (canvas_x, canvas_y).
    pdfplumber py is measured from the TOP of the display page.
    ReportLab canvas measures from the BOTTOM of the raw page.
    """
    transforms = {
        0:   lambda px, py: (px,           raw_h - py),
        90:  lambda px, py: (py,           px),
        180: lambda px, py: (raw_w - px,   py),
        270: lambda px, py: (raw_h - py,   raw_w - px),
    }
    if rotation not in transforms:
        raise ValueError(f"Unsupported rotation: {rotation}")
    return transforms[rotation]


def get_canvas_pagesize(raw_w, raw_h, rotation):
    """ReportLab canvas pagesize must match the raw MediaBox (before rotation)."""
    return (raw_w, raw_h)


# ── Step 2: Extract rooms and spaces from the PDF ─────────────────────────────

# Patterns
ROOM_NUM_RE   = re.compile(r'^[0-9]{3}[AB]$')          # e.g. 423A, 400B
UNIT_TYPE_RE  = re.compile(r'^[123][AB]$')              # e.g. 1A, 2B, 3A
SERVICE_ID_RE = re.compile(
    r'^(A|B)[0-9]{3}$'       # A401–A409, B401–B409
    r'|^(A|B)E-[0-9]$'       # AE-4, BE-4  (elevator)
    r'|^(A|B)S[12]-[0-9]$'   # AS1-4, AS2-4, BS1-4, BS2-4  (stair)
)

# Manual label lookup for service room IDs
SERVICE_LABEL_MAP = {
    'A401': 'LOBBY',       'B401': 'LOBBY',
    'A402': 'CORR',        'B402': 'CORR',
    'A403': 'ELEC',        'B403': 'ELEC',
    'A404': 'ELEC',        'B404': 'ELEC',
    'A405': 'MECH',        'B405': 'MECH',
    'A406': 'ELEC',        'B406': 'ELEC',
    'A407': 'MECH',        'B407': 'MECH',
    'A408': 'ELEV EQUIP',  'B408': 'ELEV EQUIP',
    'A409': 'TENANT STOR', 'B409': 'TENANT STOR',
    'AE-4': 'ELEV',        'BE-4': 'ELEV',
    'AS1-4': 'STAIR',      'AS2-4': 'STAIR',
    'BS1-4': 'STAIR',      'BS2-4': 'STAIR',
}


def extract_rooms(pdf_path):
    """
    Parse all words from the PDF (in display/pdfplumber space) and return a list of dicts:
        { room, type, bldg, x, y }
    where x, y are pdfplumber coordinates (origin = top-left of display page).
    """
    with pdfplumber.open(pdf_path) as pdf:
        words = pdf.pages[0].extract_words()

    items = [
        {'text': w['text'],
         'x': (w['x0'] + w['x1']) / 2,
         'y': (w['top'] + w['bottom']) / 2}
        for w in words
    ]

    results = []

    # ── Residential units ──────────────────────────────────────────────────
    for room in (i for i in items if ROOM_NUM_RE.match(i['text'])):
        rx, ry = room['x'], room['y']
        unit_label = None

        # The word "UNIT" appears ~13 pts above the room number in the same column
        for item in items:
            if item['text'] == 'UNIT' and abs(item['x'] - rx) < 40 and abs(item['y'] - ry) < 30:
                # The unit type (1A, 2B, …) is on the same line as "UNIT", slightly to its right
                for t in items:
                    if UNIT_TYPE_RE.match(t['text']) \
                            and abs(t['x'] - item['x']) < 50 \
                            and abs(t['y'] - item['y']) < 5:
                        unit_label = t['text']
                        break
                break

        bldg = 'A' if room['text'].endswith('A') else 'B'
        results.append({
            'room': room['text'],
            'type': f'UNIT {unit_label}' if unit_label else 'UNIT',
            'bldg': bldg,
            'x': rx, 'y': ry,
        })

    # ── Service / support spaces ───────────────────────────────────────────
    for sid in (i for i in items if SERVICE_ID_RE.match(i['text'])):
        bldg = 'A' if sid['text'][0] == 'A' else 'B'
        stype = SERVICE_LABEL_MAP.get(sid['text'], 'SERVICE')
        results.append({
            'room': sid['text'],
            'type': stype,
            'bldg': bldg,
            'x': sid['x'], 'y': sid['y'],
        })

    results.sort(key=lambda r: (r['bldg'], r['room']))
    return results


# ── Step 3: Sign type classification ─────────────────────────────────────────

def classify_sign(room_id, room_type):
    """Return (sign_type_label, sign_description) for a given room."""
    rt = room_type.upper()
    rid = room_id.upper()

    if 'UNIT' in rt:
        ut = rt.replace('UNIT ', '')
        beds = {'1A': '1-BED', '1B': '1-BED', '2A': '2-BED', '2B': '2-BED',
                '3A': '3-BED', '3B': '3-BED'}.get(ut, '')
        return f'{beds} UNIT SIGN', 'Suite ID Sign'
    if 'LOBBY'       in rt: return 'LOBBY SIGN',            'Directory / Lobby ID Sign'
    if 'MECH'        in rt: return 'MECHANICAL ROOM SIGN',  'Hazard / ID Sign'
    if 'ELEC'        in rt: return 'ELECTRICAL ROOM SIGN',  'Hazard / ID Sign'
    if 'STAIR'       in rt or re.match(r'^[AB]S', rid):
                             return 'STAIRWELL SIGN',        'Egress Sign'
    if 'ELEV EQUIP'  in rt: return 'ELEV EQUIPMENT SIGN',  'Elevator ID Sign'
    if 'ELEV'        in rt: return 'ELEVATOR SIGN',         'Elevator ID Sign'
    if 'TENANT STOR' in rt: return 'TENANT STORAGE SIGN',  'Room ID Sign'
    if 'CORR'        in rt: return 'CORRIDOR SIGN',         'Wayfinding Sign'
    return 'ROOM ID SIGN', 'Room ID Sign'


# ── Step 4: Build Excel sign takeoff ──────────────────────────────────────────

def build_excel(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4th Floor Sign Takeoff"

    # Styles
    HDR_FILL  = PatternFill('solid', start_color='1F3864')
    SUB_FILL  = PatternFill('solid', start_color='2E75B6')
    TOT_FILL  = PatternFill('solid', start_color='E2EFDA')
    FILL_A    = [PatternFill('solid', start_color='DDEEFF'),
                 PatternFill('solid', start_color='EEF5FF')]
    FILL_B    = [PatternFill('solid', start_color='FFF2CC'),
                 PatternFill('solid', start_color='FFFAE8')]
    thin      = Side(style='thin',   color='AAAAAA')
    med       = Side(style='medium', color='333333')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_top   = Border(left=thin, right=thin, top=med,  bottom=med)

    def hf(sz=10, bold=True, color='FFFFFF'):
        return Font(name='Arial', bold=bold, size=sz, color=color)
    def cf(bold=False, sz=9):
        return Font(name='Arial', bold=bold, size=sz, color='1A1A1A')
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'UNION AT TOWER DISTRICT — 4TH FLOOR SIGN TAKEOFF'
    ws['A1'].font = hf(14); ws['A1'].fill = HDR_FILL
    ws['A1'].alignment = ctr; ws['A1'].border = bdr
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    ws['A2'] = 'Project: Union at Tower District  |  Sheet: A1.4  |  Date: 03-28-2025  |  Architect: BVH Architecture'
    ws['A2'].font = Font(name='Arial', size=9, color='FFFFFF', italic=True)
    ws['A2'].fill = SUB_FILL; ws['A2'].alignment = ctr; ws['A2'].border = bdr
    ws.row_dimensions[2].height = 18

    ws.append([]); ws.row_dimensions[3].height = 6

    # Column headers
    COLS = ['#', 'BUILDING', 'ROOM / SPACE ID', 'UNIT TYPE', 'SIGN TYPE', 'SIGN DESCRIPTION', 'QTY', 'NOTES']
    ws.append(COLS)
    hr = ws.max_row
    ws.row_dimensions[hr].height = 22
    for c, h in enumerate(COLS, 1):
        cell = ws.cell(hr, c, h)
        cell.font = hf(); cell.fill = SUB_FILL
        cell.alignment = ctr; cell.border = bdr

    data_start = ws.max_row + 1
    summary = {}

    for idx, r in enumerate(results, 1):
        stype, sdesc = classify_sign(r['room'], r['type'])
        fill = (FILL_A if r['bldg'] == 'A' else FILL_B)[idx % 2]
        ws.append([idx, f"Building {r['bldg']}", r['room'], r['type'], stype, sdesc, 1, ''])
        row = ws.max_row
        ws.row_dimensions[row].height = 18
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.fill = fill; cell.border = bdr; cell.font = cf()
            cell.alignment = ctr if col in (1, 2, 7) else left
        summary[stype] = summary.get(stype, 0) + 1

    data_end = ws.max_row

    # Totals row
    ws.append([])
    tr = ws.max_row + 1
    for col, val in enumerate([
            'TOTAL', '', '', '', 'ALL SIGNS', '', f'=SUM(G{data_start}:G{data_end})', ''], 1):
        cell = ws.cell(tr, col, val)
        cell.fill = TOT_FILL
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = ctr; cell.border = bdr_top
    ws.row_dimensions[tr].height = 22

    # Column widths
    for i, w in enumerate([5, 14, 16, 14, 28, 26, 7, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Sign Type Summary")
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'SIGN TYPE SUMMARY — 4TH FLOOR'
    ws2['A1'].font = hf(13); ws2['A1'].fill = HDR_FILL
    ws2['A1'].alignment = ctr; ws2['A1'].border = bdr
    ws2.row_dimensions[1].height = 26
    ws2.append([])
    for c, h in enumerate(['SIGN TYPE', 'SIGN DESCRIPTION', 'COUNT', 'NOTES'], 1):
        cell = ws2.cell(3, c, h)
        cell.font = hf(); cell.fill = SUB_FILL
        cell.alignment = ctr; cell.border = bdr
    ws2.row_dimensions[3].height = 20
    ALT = [PatternFill('solid', start_color='DDEEFF'), PatternFill('solid', start_color='EEF5FF')]
    for i, (stype, count) in enumerate(sorted(summary.items())):
        _, desc = classify_sign('', stype.split()[0])
        ws2.append([stype, desc, count, ''])
        r2 = ws2.max_row
        for col in range(1, 5):
            cell = ws2.cell(r2, col)
            cell.fill = ALT[i % 2]; cell.font = cf()
            cell.alignment = ctr if col == 3 else left; cell.border = bdr
    for i, w in enumerate([32, 26, 10, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    tr2 = ws2.max_row + 1
    ws2.cell(tr2, 1, 'TOTAL'); ws2.cell(tr2, 3, f'=SUM(C4:C{ws2.max_row})')
    for col in range(1, 5):
        cell = ws2.cell(tr2, col)
        cell.fill = TOT_FILL
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = ctr; cell.border = bdr_top

    wb.save(output_path)
    print(f"✓ Excel saved → {output_path}  ({len(results)} signs, {len(summary)} types)")
    return summary


# ── Step 5: Build annotated PDF with color-coded dot markers ──────────────────

COLOR_MAP = {
    '1A': '#2196F3', '1B': '#1565C0',   # blues  – 1-bed
    '2A': '#4CAF50', '2B': '#2E7D32',   # greens – 2-bed
    '3A': '#FF9800', '3B': '#E65100',   # oranges – 3-bed
    'LOBBY':       '#9C27B0',            # purple
    'MECH':        '#F44336',            # red
    'ELEC':        '#FF5722',            # deep orange
    'STAIR':       '#607D8B',            # blue-grey
    'ELEV':        '#00BCD4',            # cyan
    'ELEV EQUIP':  '#00ACC1',            # teal
    'TENANT STOR': '#795548',            # brown
    'CORR':        '#9E9E9E',            # grey
}

LEGEND_LABELS = [
    ('1-BED UNIT (1A)', '1A'), ('1-BED UNIT (1B)', '1B'),
    ('2-BED UNIT (2A)', '2A'), ('2-BED UNIT (2B)', '2B'),
    ('3-BED UNIT (3A)', '3A'), ('3-BED UNIT (3B)', '3B'),
    ('LOBBY',           'LOBBY'),
    ('MECHANICAL ROOM', 'MECH'),
    ('ELECTRICAL ROOM', 'ELEC'),
    ('STAIRWELL',       'STAIR'),
    ('ELEVATOR',        'ELEV'),
    ('ELEV EQUIP ROOM', 'ELEV EQUIP'),
    ('TENANT STORAGE',  'TENANT STOR'),
    ('CORRIDOR',        'CORR'),
]


def room_color(room_type):
    rt = room_type.upper()
    if 'UNIT' in rt:
        key = rt.replace('UNIT ', '')
        return HexColor(COLOR_MAP.get(key, '#2196F3'))
    for k, v in COLOR_MAP.items():
        if k in rt:
            return HexColor(v)
    return HexColor('#999999')


def build_annotated_pdf(results, input_path, output_path, to_canvas, raw_w, raw_h):
    packet = io.BytesIO()
    c = rl_canvas.Canvas(packet, pagesize=(raw_w, raw_h))

    # Draw dot markers
    for r in results:
        cx, cy = to_canvas(r['x'], r['y'])
        c.setFillColor(room_color(r['type']))
        c.setStrokeColor(HexColor('#FFFFFF'))
        c.setLineWidth(1.5)
        c.circle(cx, cy, 8, stroke=1, fill=1)

    # Legend (placed at raw bottom-left = display bottom-right corner)
    leg_x, leg_y = 30, 30
    leg_w = 260
    leg_h = len(LEGEND_LABELS) * 17 + 28

    c.setFillColor(HexColor('#FFFFFF'))
    c.setStrokeColor(HexColor('#333333'))
    c.setLineWidth(1)
    c.rect(leg_x - 6, leg_y - 4, leg_w, leg_h, stroke=1, fill=1)

    c.setFillColor(HexColor('#1F3864'))
    c.rect(leg_x - 6, leg_y + leg_h - 22, leg_w, 22, stroke=0, fill=1)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont('Helvetica-Bold', 8)
    c.drawString(leg_x, leg_y + leg_h - 16, 'SIGN LEGEND — 4TH FLOOR')

    for i, (label, key) in enumerate(LEGEND_LABELS):
        iy = leg_y + leg_h - 38 - i * 17
        c.setFillColor(HexColor(COLOR_MAP[key]))
        c.setStrokeColor(HexColor('#FFFFFF'))
        c.setLineWidth(0.5)
        c.circle(leg_x + 6, iy + 4, 6, stroke=1, fill=1)
        c.setFillColor(HexColor('#1A1A1A'))
        c.setFont('Helvetica', 7)
        c.drawString(leg_x + 18, iy + 1, label)

    c.save()
    packet.seek(0)

    reader  = PdfReader(input_path)
    overlay = PdfReader(packet)
    writer  = PdfWriter()
    page    = reader.pages[0]
    page.merge_page(overlay.pages[0])
    writer.add_page(page)

    with open(output_path, 'wb') as f:
        writer.write(f)

    print(f"✓ Annotated PDF saved → {output_path}  ({len(results)} markers)")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print(f"\nProcessing: {INPUT_PDF}")

    raw_w, raw_h, rotation = get_page_info(INPUT_PDF)
    print(f"  Raw MediaBox : {raw_w} × {raw_h} pts")
    print(f"  /Rotate      : {rotation}°")

    to_canvas = make_canvas_transform(raw_w, raw_h, rotation)

    rooms = extract_rooms(INPUT_PDF)
    print(f"  Rooms found  : {len(rooms)}")

    build_excel(rooms, OUTPUT_XLS)
    build_annotated_pdf(rooms, INPUT_PDF, OUTPUT_PDF, to_canvas, raw_w, raw_h)
    print("\nDone.")
